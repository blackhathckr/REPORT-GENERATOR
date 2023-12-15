from flask import Flask,render_template,send_file,request
from reportlab.pdfgen import canvas
import openpyxl
from io import BytesIO

app=Flask(__name__)

@app.route('/',methods=['GET'])
def index():
    return render_template("index.html")

@app.route('/generate_pdf_report',methods=['POST'])
def generate_pdf_report():
    fname=request.form.get('fname')
    lname=request.form.get('lname')
    std_usn=request.form.get('usn')
    cgpa=request.form.get('cgpa')

    report_path = 'report_data.xlsx'

    workbook = openpyxl.load_workbook(report_path)

    sheet = workbook.active

    pdf_stream = BytesIO()

    pdf_canvas = canvas.Canvas(pdf_stream)
    pdf_canvas.setFont("Helvetica", 12)

    data_row = None

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        usn, name, year, sem, feedback = row
        if std_usn.lower() == str(usn).lower():
            data_row = row
            break

    if data_row:
        page_width, page_height = pdf_canvas._pagesize
        center_x = page_width / 2

        text_width = pdf_canvas.stringWidth(f"Name: {fname}", "Helvetica", 12)

        pdf_canvas.drawString(center_x - text_width / 2, 800, f"Name: {data_row[1]}")
        pdf_canvas.drawString(center_x - text_width / 2, 780, f"USN: {data_row[0]}")
        pdf_canvas.drawString(center_x - text_width / 2, 760, f"Year: {data_row[2]}")
        pdf_canvas.drawString(center_x - text_width / 2, 740, f"Sem: {data_row[3]}")
        pdf_canvas.drawString(center_x - text_width / 2, 720, f"CGPA: {cgpa}")
        pdf_canvas.drawString(center_x - text_width / 2, 700, f"Feedback: {data_row[4]}")

        pdf_canvas.save()

        pdf_stream.seek(0)

        response = send_file(pdf_stream, as_attachment=True, download_name=str(usn)+'_report.pdf')
    else:
        return render_template("error.html")

    workbook.close()

    return response

if __name__ == '__main__':
    app.run(debug=True)